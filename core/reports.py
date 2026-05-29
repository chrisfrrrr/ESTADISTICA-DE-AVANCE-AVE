from __future__ import annotations
from io import BytesIO
from pathlib import Path
from datetime import datetime, date, time
import json, math
import pandas as pd
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

DEV = 'Ing. Christian Pocol, Ingeniero Electrónico'
BLUE = '172B85'
GREEN = '00A83B'
DARK = '263238'

def _is_missing(value) -> bool:
    if value is None:
        return True
    try:
        if value is pd.NaT or pd.isna(value):
            return True
    except Exception:
        pass
    if isinstance(value, float) and math.isnan(value):
        return True
    return False

def _safe_cell(value):
    if _is_missing(value):
        return ''
    if isinstance(value, pd.Timestamp):
        if pd.isna(value):
            return ''
        try:
            if value.tzinfo is not None:
                value = value.tz_convert(None)
        except Exception:
            try: value = value.tz_localize(None)
            except Exception: pass
        return value.strftime('%Y-%m-%d %H:%M:%S')
    if isinstance(value, datetime):
        if value.tzinfo is not None: value = value.replace(tzinfo=None)
        return value.strftime('%Y-%m-%d %H:%M:%S')
    if isinstance(value, date): return value.strftime('%Y-%m-%d')
    if isinstance(value, time): return value.replace(tzinfo=None).strftime('%H:%M:%S') if value.tzinfo else value.strftime('%H:%M:%S')
    if isinstance(value, (dict, list, tuple, set)):
        try: return json.dumps(value, ensure_ascii=False, default=str)
        except Exception: return str(value)
    try:
        if hasattr(value, 'item'): value = value.item()
    except Exception: pass
    if isinstance(value, str): return value.replace('T', ' ').replace('Z', '')
    return value



def _num_series(df: pd.DataFrame | None, col: str, default: float = 0.0) -> pd.Series:
    """Devuelve una serie numérica segura aunque la columna venga como texto, vacía o no exista."""
    if df is None or col not in df.columns:
        return pd.Series([default])
    return pd.to_numeric(df[col], errors='coerce').fillna(default)

def _num_sum(df: pd.DataFrame | None, col: str, default: float = 0.0) -> float:
    return float(_num_series(df, col, default).sum())

def _num_mean(df: pd.DataFrame | None, col: str, default: float = 0.0) -> float:
    ser = _num_series(df, col, default)
    return float(ser.mean()) if len(ser) else default

def _write_df(ws, df: pd.DataFrame | None):
    if df is None or df.empty:
        ws.append(['Sin datos disponibles'])
        return
    clean = df.copy()
    clean.columns = [str(c) for c in clean.columns]
    ws.append(list(clean.columns))
    for row in clean.itertuples(index=False, name=None):
        ws.append([_safe_cell(v) for v in row])

def _format_ws(ws):
    ws.freeze_panes = 'A2'
    header_fill = PatternFill('solid', fgColor=BLUE)
    header_font = Font(color='FFFFFF', bold=True)
    thin = Side(style='thin', color='D9D9D9')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical='top', wrap_text=True)
    if ws.max_row >= 1 and ws.max_column >= 1:
        ws.auto_filter.ref = ws.dimensions
    for column_cells in ws.columns:
        values = [str(c.value) if c.value is not None else '' for c in column_cells]
        max_len = min(max([len(v) for v in values] + [10]) + 2, 60)
        ws.column_dimensions[column_cells[0].column_letter].width = max_len

def excel_bytes(summary: pd.DataFrame, submissions: pd.DataFrame, history: pd.DataFrame | None = None, followups: pd.DataFrame | None = None, modules: pd.DataFrame | None = None, module_items: pd.DataFrame | None = None, module_matrix: pd.DataFrame | None = None) -> bytes:
    bio = BytesIO()
    wb = Workbook()
    sheets = [
        ('Resumen estudiantes', summary),
        ('Entregas detalle', submissions),
        ('Historial riesgo', history),
        ('Bitacora seguimiento', followups),
        ('Modulos', modules),
        ('Items por modulo', module_items),
        ('Matriz modulos', module_matrix),
    ]
    ws = wb.active
    ws.title = sheets[0][0]
    _write_df(ws, sheets[0][1])
    for title, df in sheets[1:]:
        wsx = wb.create_sheet(title[:31])
        _write_df(wsx, df)
    if summary is not None and not summary.empty:
        ws_kpi = wb.create_sheet('Indicadores ejecutivos')
        counts = summary['riesgo_integral'].value_counts().to_dict() if 'riesgo_integral' in summary else {}
        kpis = [
            ['Indicador','Valor'],
            ['Total estudiantes', len(summary)],
            ['Riesgo alto', counts.get('Alto',0)],
            ['Riesgo medio', counts.get('Medio',0)],
            ['Riesgo bajo', counts.get('Bajo',0)],
            ['Avance promedio', round(_num_mean(summary, 'porcentaje_avance'), 2)],
            ['Horas promedio', round(_num_mean(summary, 'tiempo_total_horas'), 2)],
            ['Pendientes actuales', int(_num_sum(summary, 'pendientes_actuales') if 'pendientes_actuales' in summary.columns else _num_sum(summary, 'pendientes'))],
            ['Pendientes futuros', int(_num_sum(summary, 'pendientes_futuros'))],
            ['Pendientes totales', int(_num_sum(summary, 'pendientes_total'))],
            ['Atrasadas totales', int(_num_sum(summary, 'atrasadas'))],
        ]
        for r in kpis: ws_kpi.append(r)
    for wsx in wb.worksheets:
        _format_ws(wsx)
    wb.save(bio)
    return bio.getvalue()

def _watermark(canvas, doc):
    canvas.saveState()
    canvas.setFont('Helvetica-Bold', 42)
    canvas.setFillColor(colors.Color(0.85, 0.85, 0.85, alpha=0.23))
    canvas.translate(5.5 * inch, 4.2 * inch)
    canvas.rotate(35)
    canvas.drawCentredString(0, 0, 'AVE - UVG')
    canvas.restoreState()
    canvas.saveState()
    canvas.setFont('Helvetica', 8)
    canvas.setFillColor(colors.grey)
    canvas.drawString(0.35 * inch, 0.25 * inch, f'Desarrollador: {DEV}')
    canvas.drawRightString(10.65 * inch, 0.25 * inch, f'Página {doc.page}')
    canvas.restoreState()

def _styles():
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name='SmallAVE', fontSize=8, leading=10))
    styles.add(ParagraphStyle(name='TinyAVE', fontSize=6.5, leading=8))
    styles.add(ParagraphStyle(name='TitleAVE', fontSize=18, leading=22, alignment=1, textColor=colors.HexColor('#172B85')))
    styles.add(ParagraphStyle(name='SectionAVE', fontSize=12, leading=14, textColor=colors.HexColor('#172B85'), spaceBefore=10, spaceAfter=6))
    return styles

def _header(story, styles, title, course_name, section_name, generated_by, analysis_date, logo_ave, logo_uvg):
    header_data = []
    header_data.append(Image(logo_ave, width=1.25*inch, height=0.62*inch) if Path(logo_ave).exists() else '')
    header_data.append(Paragraph(title, styles['TitleAVE']))
    header_data.append(Image(logo_uvg, width=1.1*inch, height=0.62*inch) if Path(logo_uvg).exists() else '')
    ht = Table([header_data], colWidths=[1.4*inch, 7.2*inch, 1.4*inch])
    ht.setStyle(TableStyle([('VALIGN',(0,0),(-1,-1),'MIDDLE')]))
    story += [ht, Spacer(1, 10)]
    story.append(Paragraph(f'<b>Curso:</b> {course_name}<br/><b>Sección:</b> {section_name}<br/><b>Fecha de análisis:</b> {analysis_date}<br/><b>Generado por:</b> {generated_by or "No especificado"}<br/><b>Desarrollador:</b> {DEV}', styles['SmallAVE']))
    story.append(Spacer(1, 10))

def pdf_bytes(summary: pd.DataFrame, course_name: str, section_name: str, generated_by: str, analysis_date: str, logo_ave='assets/logo_ave.png', logo_uvg='assets/logo_uvg.png') -> bytes:
    bio = BytesIO()
    doc = SimpleDocTemplate(bio, pagesize=landscape(letter), rightMargin=28, leftMargin=28, topMargin=28, bottomMargin=28)
    styles = _styles()
    story = []
    _header(story, styles, 'Informe Ejecutivo de Seguimiento Académico AVE', course_name, section_name, generated_by, analysis_date, logo_ave, logo_uvg)
    total = len(summary) if summary is not None else 0
    counts = summary['riesgo_integral'].value_counts().to_dict() if total and 'riesgo_integral' in summary.columns else {}
    kpi = [['Total estudiantes','Riesgo bajo','Riesgo medio','Riesgo alto','Prom. avance','Pendientes actuales','Pendientes futuros','Atrasadas'],[total, counts.get('Bajo',0), counts.get('Medio',0), counts.get('Alto',0), f"{_num_mean(summary, 'porcentaje_avance'):.1f}%" if total else '0%', int(_num_sum(summary, 'pendientes_actuales') if 'pendientes_actuales' in summary.columns else _num_sum(summary, 'pendientes')) if total else 0, int(_num_sum(summary, 'pendientes_futuros')) if total else 0, int(_num_sum(summary, 'atrasadas')) if total else 0]]
    kt = Table(kpi, colWidths=[1.15*inch]*8)
    kt.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),colors.HexColor('#172B85')),('TEXTCOLOR',(0,0),(-1,0),colors.white),('GRID',(0,0),(-1,-1),0.25,colors.grey),('ALIGN',(0,0),(-1,-1),'CENTER'),('FONTSIZE',(0,0),(-1,-1),8)]))
    story += [kt, Spacer(1, 10)]
    story.append(Paragraph('<b>Interpretación ejecutiva:</b> el riesgo integral prioriza desconexión, déficit de horas, entregas pendientes actuales, pendientes futuros, entregas atrasadas, bajo avance y reincidencia operativa. El listado siguiente incluye todos los casos de riesgo Alto y Medio para intervención del asesor.', styles['SmallAVE']))
    story.append(Spacer(1, 8))
    if summary is None or summary.empty:
        story.append(Paragraph('No hay estudiantes para mostrar.', styles['SmallAVE']))
    else:
        cols = ['estudiante','correo','horas_sin_actividad','riesgo_desconexion','pendientes_actuales','pendientes_futuros','atrasadas','porcentaje_avance','puntaje_riesgo','riesgo_integral','segmento_ave','accion_recomendada']
        view = summary.copy()
        for c in cols:
            if c not in view.columns: view[c] = ''
        order = {'Alto':0,'Medio':1,'Bajo':2}
        view['_ord'] = view['riesgo_integral'].map(order).fillna(3)
        view = view[view['riesgo_integral'].isin(['Alto','Medio'])].sort_values(['_ord','puntaje_riesgo','horas_sin_actividad'], ascending=[True,False,False])
        if view.empty:
            view = summary.sort_values(['puntaje_riesgo','horas_sin_actividad'], ascending=[False,False])
        story.append(Paragraph(f'<b>Casos listados:</b> {len(view)} estudiantes. Riesgo alto: {counts.get("Alto",0)}. Riesgo medio: {counts.get("Medio",0)}.', styles['SmallAVE']))
        data = [['Estudiante','Correo','Hrs sin act.','Riesgo conexión','Pend. actuales','Pend. futuros','Atr.','Avance actual %','Puntaje','Riesgo','Segmento','Acción']] + view[cols].fillna('').astype(str).values.tolist()
        table = Table(data, repeatRows=1, colWidths=[1.28*inch,1.25*inch,0.58*inch,0.72*inch,0.48*inch,0.48*inch,0.42*inch,0.50*inch,0.48*inch,0.52*inch,1.0*inch,1.05*inch])
        table.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),colors.HexColor('#6c757d')),('TEXTCOLOR',(0,0),(-1,0),colors.white),('GRID',(0,0),(-1,-1),0.22,colors.lightgrey),('FONTSIZE',(0,0),(-1,-1),6),('VALIGN',(0,0),(-1,-1),'TOP'),('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white, colors.HexColor('#F7F9FB')])]))
        story.append(table)
    story.append(Spacer(1, 8))
    story.append(Paragraph('Nota: los datos dependen de los permisos del token y de los registros disponibles en Canvas. Se recomienda complementar la lectura con la bitácora de seguimiento del asesor.', styles['SmallAVE']))
    doc.build(story, onFirstPage=_watermark, onLaterPages=_watermark)
    return bio.getvalue()

def individual_pdf_bytes(student_row: dict, submissions: pd.DataFrame, followups: pd.DataFrame, course_name: str, section_name: str, generated_by: str, analysis_date: str, logo_ave='assets/logo_ave.png', logo_uvg='assets/logo_uvg.png') -> bytes:
    bio = BytesIO()
    doc = SimpleDocTemplate(bio, pagesize=letter, rightMargin=36, leftMargin=36, topMargin=30, bottomMargin=30)
    styles = _styles()
    story = []
    _header(story, styles, 'Ficha Individual de Seguimiento AVE', course_name, section_name, generated_by, analysis_date, logo_ave, logo_uvg)
    name = student_row.get('estudiante') or student_row.get('nombre') or 'Estudiante'
    story.append(Paragraph(f'<b>Estudiante:</b> {name}<br/><b>Correo:</b> {student_row.get("correo", "")}<br/><b>Riesgo integral:</b> {student_row.get("riesgo_integral", "")}<br/><b>Segmento AVE:</b> {student_row.get("segmento_ave", "")}<br/><b>Acción recomendada:</b> {student_row.get("accion_recomendada", "")}', styles['SmallAVE']))
    story.append(Spacer(1, 10))
    kpi = [['Horas sin actividad','Horas acumuladas','Horas esperadas','Déficit','Pendientes actuales','Pendientes futuros','Atrasadas','Avance %','Puntaje'],[student_row.get('horas_sin_actividad',''), student_row.get('tiempo_total_horas',''), student_row.get('horas_esperadas',''), student_row.get('deficit_horas',''), student_row.get('pendientes_actuales', student_row.get('pendientes','')), student_row.get('pendientes_futuros',''), student_row.get('atrasadas',''), student_row.get('porcentaje_avance',''), student_row.get('puntaje_riesgo','')]]
    kt = Table(kpi, colWidths=[0.78*inch]*9)
    kt.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),colors.HexColor('#172B85')),('TEXTCOLOR',(0,0),(-1,0),colors.white),('GRID',(0,0),(-1,-1),0.25,colors.grey),('FONTSIZE',(0,0),(-1,-1),7),('ALIGN',(0,0),(-1,-1),'CENTER')]))
    story += [kt, Spacer(1, 12)]
    story.append(Paragraph('Entregas del estudiante', styles['SectionAVE']))
    uid = student_row.get('user_id')
    sub = submissions[submissions['user_id'].astype(str)==str(uid)].copy() if submissions is not None and not submissions.empty and uid is not None else pd.DataFrame()
    if sub.empty:
        story.append(Paragraph('Sin detalle de entregas disponible para este estudiante.', styles['SmallAVE']))
    else:
        cols = [c for c in ['actividad','fecha_entrega','submitted_at','workflow_state','missing','late','score','puntos'] if c in sub.columns]
        data = [cols] + sub[cols].fillna('').astype(str).head(80).values.tolist()
        t = Table(data, repeatRows=1, colWidths=[2.1*inch,0.9*inch,0.9*inch,0.8*inch,0.45*inch,0.45*inch,0.45*inch,0.45*inch][:len(cols)])
        t.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),colors.HexColor('#6c757d')),('TEXTCOLOR',(0,0),(-1,0),colors.white),('GRID',(0,0),(-1,-1),0.22,colors.lightgrey),('FONTSIZE',(0,0),(-1,-1),6.2),('VALIGN',(0,0),(-1,-1),'TOP')]))
        story.append(t)
    story.append(Spacer(1, 10))
    story.append(Paragraph('Bitácora de seguimiento', styles['SectionAVE']))
    bit = followups[followups['user_id'].astype(str)==str(uid)].copy() if followups is not None and not followups.empty and uid is not None else pd.DataFrame()
    if bit.empty:
        story.append(Paragraph('Sin registros de seguimiento en la bitácora local.', styles['SmallAVE']))
    else:
        cols = [c for c in ['created_at','medio','motivo','resultado','proxima_accion','fecha_proxima_accion','observaciones','registrado_por'] if c in bit.columns]
        data = [cols] + bit[cols].fillna('').astype(str).head(30).values.tolist()
        t = Table(data, repeatRows=1, colWidths=[0.85*inch,0.7*inch,0.85*inch,0.9*inch,0.9*inch,0.75*inch,1.4*inch,0.8*inch][:len(cols)])
        t.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),colors.HexColor('#6c757d')),('TEXTCOLOR',(0,0),(-1,0),colors.white),('GRID',(0,0),(-1,-1),0.22,colors.lightgrey),('FONTSIZE',(0,0),(-1,-1),6.2),('VALIGN',(0,0),(-1,-1),'TOP')]))
        story.append(t)
    doc.build(story, onFirstPage=_watermark, onLaterPages=_watermark)
    return bio.getvalue()

# ===== PRO 4.0: PDF ejecutivo con gráficas institucionales =====
def _risk_pie_drawing(summary: pd.DataFrame):
    from reportlab.graphics.shapes import Drawing, String
    from reportlab.graphics.charts.piecharts import Pie
    d = Drawing(300, 180)
    if summary is None or summary.empty or 'riesgo_integral' not in summary.columns:
        d.add(String(20, 80, 'Sin datos de riesgo', fontSize=10))
        return d
    counts = summary[summary['riesgo_integral'].isin(['Bajo','Medio','Alto'])]['riesgo_integral'].value_counts()
    labels = [f'{k} ({int(v)})' for k, v in counts.items()]
    values = [int(v) for v in counts.values]
    if not values:
        d.add(String(20, 80, 'Sin datos de riesgo', fontSize=10))
        return d
    pie = Pie()
    pie.x = 35; pie.y = 20; pie.width = 130; pie.height = 130
    pie.data = values
    pie.labels = labels
    palette = {'Bajo': colors.HexColor('#00A83B'), 'Medio': colors.HexColor('#F2C94C'), 'Alto': colors.HexColor('#D64545')}
    for i, key in enumerate(counts.index):
        pie.slices[i].fillColor = palette.get(key, colors.HexColor('#172B85'))
    d.add(String(15, 160, 'Distribución del riesgo integral', fontSize=11, fillColor=colors.HexColor('#172B85')))
    d.add(pie)
    return d


def _course_bar_drawing(summary: pd.DataFrame):
    from reportlab.graphics.shapes import Drawing, String
    from reportlab.graphics.charts.barcharts import VerticalBarChart
    d = Drawing(430, 180)
    if summary is None or summary.empty or 'curso_general' not in summary.columns:
        d.add(String(20, 80, 'Sin datos por curso', fontSize=10))
        return d
    g = summary.groupby('curso_general').agg(
        total=('user_id','count'),
        alto=('riesgo_integral', lambda s: int((s == 'Alto').sum()) if 'riesgo_integral' in summary.columns else 0)
    ).reset_index().head(8)
    if g.empty:
        d.add(String(20, 80, 'Sin datos por curso', fontSize=10))
        return d
    chart = VerticalBarChart()
    chart.x = 35; chart.y = 35; chart.height = 105; chart.width = 360
    chart.data = [g['total'].astype(int).tolist(), g['alto'].astype(int).tolist()]
    chart.categoryAxis.categoryNames = [str(x)[:14] for x in g['curso_general'].tolist()]
    chart.categoryAxis.labels.angle = 30
    chart.categoryAxis.labels.fontSize = 6
    chart.valueAxis.valueMin = 0
    chart.bars[0].fillColor = colors.HexColor('#172B85')
    chart.bars[1].fillColor = colors.HexColor('#D64545')
    d.add(String(15, 160, 'Estudiantes y riesgo alto por curso', fontSize=11, fillColor=colors.HexColor('#172B85')))
    d.add(chart)
    return d


_pdf_bytes_old = pdf_bytes

def pdf_bytes(summary: pd.DataFrame, course_name: str, section_name: str, generated_by: str, analysis_date: str, logo_ave='assets/logo_ave.png', logo_uvg='assets/logo_uvg.png') -> bytes:
    """PDF ejecutivo AVE/UVG con gráficas y tablas de decisión.

    Reemplaza la versión previa conservando la misma firma para compatibilidad.
    """
    bio = BytesIO()
    doc = SimpleDocTemplate(bio, pagesize=landscape(letter), rightMargin=28, leftMargin=28, topMargin=28, bottomMargin=28)
    styles = _styles()
    story = []
    _header(story, styles, 'Informe Ejecutivo Global de Seguimiento Académico AVE', course_name, section_name, generated_by, analysis_date, logo_ave, logo_uvg)

    total = len(summary) if summary is not None else 0
    if summary is None:
        summary = pd.DataFrame()
    counts = summary['riesgo_integral'].value_counts().to_dict() if total and 'riesgo_integral' in summary.columns else {}
    finalizados = int(summary.get('estado_finalizacion', pd.Series(dtype=str)).eq('Finalizado o al día por entregas').sum()) if total and 'estado_finalizacion' in summary.columns else 0
    no_reg = int(summary.get('clasificacion_registro', pd.Series(dtype=str)).astype(str).str.contains('no registrado|incompletos|Pendiente', case=False, na=False).sum()) if total and 'clasificacion_registro' in summary.columns else 0
    avance_prom = f"{_num_mean(summary, 'porcentaje_avance'):.1f}%" if total else '0%'
    curso_count = summary['curso_general'].nunique() if total and 'curso_general' in summary.columns else 1

    kpi = [[
        'Cursos generales','Estudiantes detectados','Riesgo alto','Riesgo medio','Finalizados/al día','No registrados','Avance prom. actual','Pendientes actuales'
    ],[
        curso_count, total, counts.get('Alto',0), counts.get('Medio',0), finalizados, no_reg, avance_prom,
        int(_num_sum(summary, 'pendientes_actuales') if 'pendientes_actuales' in summary.columns else _num_sum(summary, 'pendientes')) if total else 0
    ]]
    kt = Table(kpi, colWidths=[1.25*inch]*8)
    kt.setStyle(TableStyle([
        ('BACKGROUND',(0,0),(-1,0),colors.HexColor('#172B85')),('TEXTCOLOR',(0,0),(-1,0),colors.white),
        ('BACKGROUND',(0,1),(-1,1),colors.HexColor('#F7F9FB')),('GRID',(0,0),(-1,-1),0.25,colors.grey),
        ('ALIGN',(0,0),(-1,-1),'CENTER'),('FONTSIZE',(0,0),(-1,-1),8),('FONTNAME',(0,0),(-1,0),'Helvetica-Bold')
    ]))
    story += [kt, Spacer(1, 10)]
    story.append(Paragraph('<b>Resumen ejecutivo:</b> el informe consolida la actividad de Canvas, entregables, estudiantes no registrados o pendientes, y riesgo ajustado. El riesgo por desconexión se reduce cuando el estudiante ya finalizó o se encuentra al día con sus entregas.', styles['SmallAVE']))
    story.append(Spacer(1, 8))

    # Graficas en PDF
    graph_table = Table([[_risk_pie_drawing(summary), _course_bar_drawing(summary)]], colWidths=[3.7*inch, 6.2*inch])
    graph_table.setStyle(TableStyle([('VALIGN',(0,0),(-1,-1),'TOP')]))
    story += [graph_table, Spacer(1, 8)]

    if total and 'curso_general' in summary.columns:
        story.append(Paragraph('Resumen global por curso', styles['SectionAVE']))
        tmp = summary.copy()
        # Asegurar que los campos agregados sean numéricos, aunque Canvas o el CSV los devuelvan como texto.
        for _c in ['pendientes_actuales', 'pendientes', 'porcentaje_avance']:
            if _c in tmp.columns:
                tmp[_c] = pd.to_numeric(tmp[_c], errors='coerce').fillna(0)
        if 'pendientes_actuales' not in tmp.columns:
            tmp['pendientes_actuales'] = tmp['pendientes'] if 'pendientes' in tmp.columns else 0
        if 'porcentaje_avance' not in tmp.columns:
            tmp['porcentaje_avance'] = 0
        g = tmp.groupby('curso_general').agg(
            detectados=('user_id','count'),
            activos=('clasificacion_registro', lambda s: int((s == 'Activo analizado').sum()) if 'clasificacion_registro' in tmp else len(s)),
            finalizados=('estado_finalizacion', lambda s: int((s == 'Finalizado o al día por entregas').sum()) if 'estado_finalizacion' in tmp else 0),
            alto=('riesgo_integral', lambda s: int((s == 'Alto').sum()) if 'riesgo_integral' in tmp else 0),
            medio=('riesgo_integral', lambda s: int((s == 'Medio').sum()) if 'riesgo_integral' in tmp else 0),
            pendientes=('pendientes_actuales','sum'),
            avance=('porcentaje_avance','mean'),
        ).reset_index().head(12)
        g['pendientes'] = pd.to_numeric(g['pendientes'], errors='coerce').fillna(0).astype(int)
        g['avance'] = pd.to_numeric(g['avance'], errors='coerce').fillna(0).round(1).astype(str) + '%'
        data = [['Curso','Detectados','Activos','Finalizados','R. alto','R. medio','Pendientes','Avance prom.']] + g.fillna('').astype(str).values.tolist()
        t = Table(data, repeatRows=1, colWidths=[2.05*inch,0.7*inch,0.58*inch,0.72*inch,0.55*inch,0.55*inch,0.7*inch,0.82*inch])
        t.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),colors.HexColor('#00A83B')),('TEXTCOLOR',(0,0),(-1,0),colors.white),('GRID',(0,0),(-1,-1),0.22,colors.lightgrey),('FONTSIZE',(0,0),(-1,-1),7),('VALIGN',(0,0),(-1,-1),'TOP'),('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white, colors.HexColor('#F7F9FB')])]))
        story.append(t)
        story.append(Spacer(1, 8))

    story.append(Paragraph('Casos prioritarios para seguimiento', styles['SectionAVE']))
    if summary.empty:
        story.append(Paragraph('No hay estudiantes para mostrar.', styles['SmallAVE']))
    else:
        cols = ['curso_general','estudiante','correo','clasificacion_registro','estado_finalizacion','horas_sin_actividad','pendientes_actuales','atrasadas','porcentaje_avance','porcentaje_avance_curso','riesgo_integral','accion_recomendada']
        view = summary.copy()
        for c in cols:
            if c not in view.columns:
                view[c] = ''
        if 'puntaje_riesgo' in view.columns:
            view = view.sort_values(['riesgo_integral','puntaje_riesgo'], ascending=[True,False])
        view = view[(view['riesgo_integral'].isin(['Alto','Medio'])) | (~view['clasificacion_registro'].eq('Activo analizado'))].head(35)
        if view.empty:
            view = summary.head(25)
        data = [['Curso','Estudiante','Correo','Registro','Estado','Hrs sin act.','Pend.','Atr.','Av. act.','Av. curso','Riesgo','Acción']] + view[cols].fillna('').astype(str).values.tolist()
        table = Table(data, repeatRows=1, colWidths=[0.95*inch,1.15*inch,1.05*inch,0.85*inch,0.85*inch,0.48*inch,0.42*inch,0.35*inch,0.42*inch,0.45*inch,0.45*inch,1.25*inch])
        table.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),colors.HexColor('#6c757d')),('TEXTCOLOR',(0,0),(-1,0),colors.white),('GRID',(0,0),(-1,-1),0.22,colors.lightgrey),('FONTSIZE',(0,0),(-1,-1),5.8),('VALIGN',(0,0),(-1,-1),'TOP'),('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white, colors.HexColor('#F7F9FB')])]))
        story.append(table)
    story.append(Spacer(1, 8))
    story.append(Paragraph('Nota: los datos dependen de los permisos del token y de los registros disponibles en Canvas. Este reporte debe utilizarse como insumo de seguimiento y no como único criterio de decisión académica.', styles['SmallAVE']))
    doc.build(story, onFirstPage=_watermark, onLaterPages=_watermark)
    return bio.getvalue()

# ===== PRO 4.0 FIX4: Informe ejecutivo premium con resumen general y gráficas =====
def _text_value(v, default=''):
    if _is_missing(v):
        return default
    return str(v)


def _as_num(df: pd.DataFrame, col: str, default=0):
    if df is None or col not in df.columns:
        return pd.Series([default])
    return pd.to_numeric(df[col], errors='coerce').fillna(default)


def _course_label_from_summary(summary: pd.DataFrame, fallback: str) -> str:
    if summary is not None and not summary.empty and 'curso_general' in summary.columns:
        cursos = sorted([str(x).strip() for x in summary['curso_general'].dropna().unique() if str(x).strip()])
        if len(cursos) == 1:
            return cursos[0]
        if 1 < len(cursos) <= 4:
            return 'Análisis global: ' + ', '.join(cursos)
        if len(cursos) > 4:
            return f'Análisis global de {len(cursos)} cursos generales'
    return fallback or 'Informe global AVE'


def _risk_counts(summary: pd.DataFrame) -> dict:
    if summary is None or summary.empty or 'riesgo_integral' not in summary.columns:
        return {'Bajo': 0, 'Medio': 0, 'Alto': 0}
    vc = summary['riesgo_integral'].fillna('Sin clasificar').astype(str).value_counts().to_dict()
    return {'Bajo': int(vc.get('Bajo', 0)), 'Medio': int(vc.get('Medio', 0)), 'Alto': int(vc.get('Alto', 0)), 'Sin clasificar': int(vc.get('Sin clasificar', 0))}


def _p(text, style, max_len=None):
    txt = _text_value(text)
    if max_len and len(txt) > max_len:
        txt = txt[:max_len-1] + '…'
    # Escapes básicos para ReportLab Paragraph
    txt = txt.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
    return Paragraph(txt, style)


def _executive_page(canvas, doc):
    canvas.saveState()
    w, h = landscape(letter)
    canvas.setFillColor(colors.HexColor('#172B85'))
    canvas.rect(0, h-0.18*inch, w, 0.18*inch, stroke=0, fill=1)
    canvas.setFillColor(colors.HexColor('#00A83B'))
    canvas.rect(0, h-0.25*inch, w, 0.07*inch, stroke=0, fill=1)
    canvas.setFont('Helvetica', 8)
    canvas.setFillColor(colors.HexColor('#6c757d'))
    canvas.drawString(0.35 * inch, 0.24 * inch, f'Desarrollador: {DEV}')
    canvas.drawRightString(w - 0.35 * inch, 0.24 * inch, f'Página {doc.page}')
    canvas.restoreState()


def _premium_styles():
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name='CoverTitlePremium', fontSize=26, leading=30, alignment=1, textColor=colors.HexColor('#172B85'), fontName='Helvetica-Bold'))
    styles.add(ParagraphStyle(name='CoverSubtitlePremium', fontSize=12.5, leading=16, alignment=1, textColor=colors.HexColor('#263238')))
    styles.add(ParagraphStyle(name='SectionPremium', fontSize=14, leading=17, textColor=colors.HexColor('#172B85'), fontName='Helvetica-Bold', spaceBefore=9, spaceAfter=7))
    styles.add(ParagraphStyle(name='BodyPremium', fontSize=8.7, leading=11.5, textColor=colors.HexColor('#263238')))
    styles.add(ParagraphStyle(name='SmallPremium', fontSize=7.2, leading=9.2, textColor=colors.HexColor('#263238')))
    styles.add(ParagraphStyle(name='TinyPremium', fontSize=6.4, leading=7.8, textColor=colors.HexColor('#263238')))
    styles.add(ParagraphStyle(name='KpiTitle', fontSize=7.2, leading=8.5, alignment=1, textColor=colors.white, fontName='Helvetica-Bold'))
    styles.add(ParagraphStyle(name='KpiValue', fontSize=14, leading=16, alignment=1, textColor=colors.HexColor('#172B85'), fontName='Helvetica-Bold'))
    return styles


def _kpi_cards(data, styles):
    # data: [(title, value, color_hex), ...]
    row_titles, row_values = [], []
    for title, val, bg in data:
        row_titles.append(Paragraph(str(title), styles['KpiTitle']))
        row_values.append(Paragraph(str(val), styles['KpiValue']))
    t = Table([row_titles, row_values], colWidths=[1.18*inch]*len(data), rowHeights=[0.34*inch, 0.46*inch])
    style = [
        ('GRID',(0,0),(-1,-1),0.35,colors.HexColor('#D9E2F3')),
        ('BACKGROUND',(0,1),(-1,1),colors.HexColor('#F7F9FB')),
        ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
        ('ALIGN',(0,0),(-1,-1),'CENTER'),
    ]
    for i, (_, _, bg) in enumerate(data):
        style.append(('BACKGROUND',(i,0),(i,0),colors.HexColor(bg)))
    t.setStyle(TableStyle(style))
    return t


def _risk_pie_premium(summary: pd.DataFrame):
    from reportlab.graphics.shapes import Drawing, String, Rect
    from reportlab.graphics.charts.piecharts import Pie
    d = Drawing(330, 210)
    d.add(String(12, 193, 'Distribución del riesgo integral', fontSize=11, fillColor=colors.HexColor('#172B85'), fontName='Helvetica-Bold'))
    counts = _risk_counts(summary)
    values = [counts.get('Bajo', 0), counts.get('Medio', 0), counts.get('Alto', 0)]
    labels = [f'Bajo ({values[0]})', f'Medio ({values[1]})', f'Alto ({values[2]})']
    if sum(values) == 0:
        d.add(String(45, 100, 'Sin datos disponibles', fontSize=9, fillColor=colors.grey))
        return d
    pie = Pie()
    pie.x = 34; pie.y = 32; pie.width = 125; pie.height = 125
    pie.data = values
    pie.labels = labels
    pie.sideLabels = True
    pie.slices[0].fillColor = colors.HexColor('#00A83B')
    pie.slices[1].fillColor = colors.HexColor('#F2C94C')
    pie.slices[2].fillColor = colors.HexColor('#D64545')
    d.add(pie)
    total = max(sum(values), 1)
    legend = [('Bajo', '#00A83B', values[0]), ('Medio', '#F2C94C', values[1]), ('Alto', '#D64545', values[2])]
    y = 132
    for name, col, val in legend:
        d.add(Rect(205, y-6, 10, 10, fillColor=colors.HexColor(col), strokeColor=None))
        d.add(String(220, y-4, f'{name}: {val} estudiantes ({val/total:.1%})', fontSize=7.5, fillColor=colors.HexColor('#263238')))
        y -= 20
    return d


def _course_risk_chart_premium(summary: pd.DataFrame):
    from reportlab.graphics.shapes import Drawing, String
    from reportlab.graphics.charts.barcharts import VerticalBarChart
    d = Drawing(455, 210)
    d.add(String(12, 193, 'Riesgo alto y medio por curso general', fontSize=11, fillColor=colors.HexColor('#172B85'), fontName='Helvetica-Bold'))
    if summary is None or summary.empty or 'curso_general' not in summary.columns:
        d.add(String(45, 100, 'Sin datos por curso', fontSize=9, fillColor=colors.grey))
        return d
    tmp = summary.copy()
    if 'riesgo_integral' not in tmp.columns:
        tmp['riesgo_integral'] = 'Sin clasificar'
    g = tmp.groupby('curso_general').agg(
        alto=('riesgo_integral', lambda s: int((s == 'Alto').sum())),
        medio=('riesgo_integral', lambda s: int((s == 'Medio').sum())),
        total=('riesgo_integral', 'count')
    ).reset_index()
    g = g.sort_values(['alto','medio','total'], ascending=False).head(8)
    if g.empty:
        d.add(String(45, 100, 'Sin datos por curso', fontSize=9, fillColor=colors.grey))
        return d
    chart = VerticalBarChart()
    chart.x = 36; chart.y = 38; chart.height = 120; chart.width = 360
    chart.data = [g['alto'].astype(int).tolist(), g['medio'].astype(int).tolist()]
    names = [str(x).replace('Fundamentos de la Comunicación','Fund. Comunicación').replace('Iniciando tu Experiencia Virtual','Experiencia Virtual')[:18] for x in g['curso_general'].tolist()]
    chart.categoryAxis.categoryNames = names
    chart.categoryAxis.labels.angle = 25
    chart.categoryAxis.labels.fontSize = 6
    chart.valueAxis.valueMin = 0
    chart.bars[0].fillColor = colors.HexColor('#D64545')
    chart.bars[1].fillColor = colors.HexColor('#F2C94C')
    d.add(chart)
    d.add(String(360, 155, 'Rojo: riesgo alto', fontSize=7, fillColor=colors.HexColor('#D64545')))
    d.add(String(360, 142, 'Amarillo: riesgo medio', fontSize=7, fillColor=colors.HexColor('#9A7400')))
    return d


def _completion_chart_premium(summary: pd.DataFrame):
    from reportlab.graphics.shapes import Drawing, String
    from reportlab.graphics.charts.barcharts import HorizontalBarChart
    d = Drawing(455, 205)
    d.add(String(12, 188, 'Entregables y estado de avance', fontSize=11, fillColor=colors.HexColor('#172B85'), fontName='Helvetica-Bold'))
    if summary is None or summary.empty:
        d.add(String(45, 100, 'Sin datos de entregables', fontSize=9, fillColor=colors.grey))
        return d
    final_col = summary['estado_finalizacion'].astype(str) if 'estado_finalizacion' in summary.columns else pd.Series(['No determinado']*len(summary))
    finalizados = int(final_col.eq('Finalizado o al día por entregas').sum())
    en_seguimiento = int(len(summary) - finalizados)
    no_reg = int(summary['clasificacion_registro'].astype(str).str.contains('no registrado|pendiente|incompleto', case=False, na=False).sum()) if 'clasificacion_registro' in summary.columns else 0
    pendientes = int(_num_sum(summary, 'pendientes_actuales') if 'pendientes_actuales' in summary.columns else _num_sum(summary, 'pendientes'))
    labels = ['Finalizados/al día', 'En seguimiento', 'Registros pendientes', 'Pendientes actuales']
    values = [finalizados, en_seguimiento, no_reg, pendientes]
    chart = HorizontalBarChart()
    chart.x = 135; chart.y = 45; chart.height = 108; chart.width = 260
    chart.data = [values]
    chart.categoryAxis.categoryNames = labels
    chart.categoryAxis.labels.fontSize = 7
    chart.valueAxis.valueMin = 0
    chart.bars[0].fillColor = colors.HexColor('#172B85')
    d.add(chart)
    return d


def _executive_findings(summary: pd.DataFrame, styles) -> Table:
    total = len(summary) if summary is not None else 0
    counts = _risk_counts(summary)
    high = counts.get('Alto', 0)
    med = counts.get('Medio', 0)
    pct_high = (high / total * 100) if total else 0
    finalizados = int(summary['estado_finalizacion'].astype(str).eq('Finalizado o al día por entregas').sum()) if total and 'estado_finalizacion' in summary.columns else 0
    no_reg = int(summary['clasificacion_registro'].astype(str).str.contains('no registrado|pendiente|incompleto', case=False, na=False).sum()) if total and 'clasificacion_registro' in summary.columns else 0
    pendientes = int(_num_sum(summary, 'pendientes_actuales') if total and 'pendientes_actuales' in summary.columns else _num_sum(summary, 'pendientes')) if total else 0
    course_top = 'No determinado'
    if total and 'curso_general' in summary.columns and 'riesgo_integral' in summary.columns:
        g = summary.groupby('curso_general')['riesgo_integral'].apply(lambda s: int((s == 'Alto').sum())).sort_values(ascending=False)
        if not g.empty and int(g.iloc[0]) > 0:
            course_top = f'{g.index[0]} ({int(g.iloc[0])} casos alto)'
    findings = [
        ['Hallazgo ejecutivo', 'Lectura operativa'],
        ['Riesgo alto', f'{high} de {total} estudiantes ({pct_high:.1f}%) requieren priorización inmediata.'],
        ['Riesgo medio', f'{med} estudiantes requieren prevención y seguimiento antes de convertirse en alerta crítica.'],
        ['Entregables', f'{finalizados} estudiantes aparecen finalizados o al día; este dato evita clasificarlos solo por baja conexión.'],
        ['Pendientes actuales', f'Se detectan {pendientes} pendientes actuales acumulados en los estudiantes analizados.'],
        ['Registros incompletos', f'{no_reg} estudiantes presentan información pendiente, no registrada o incompleta y deben verificarse en Canvas.'],
        ['Curso con mayor atención', course_top],
    ]
    t = Table([[ _p(c, styles['SmallPremium'], 36), _p(v, styles['SmallPremium'], 150)] for c, v in findings], colWidths=[1.65*inch, 7.55*inch])
    t.setStyle(TableStyle([
        ('BACKGROUND',(0,0),(-1,0),colors.HexColor('#172B85')),('TEXTCOLOR',(0,0),(-1,0),colors.white),
        ('GRID',(0,0),(-1,-1),0.25,colors.HexColor('#D9E2F3')),
        ('VALIGN',(0,0),(-1,-1),'TOP'),
        ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white, colors.HexColor('#F7F9FB')]),
        ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold')
    ]))
    return t


def _recommendations(summary: pd.DataFrame, styles):
    recs = [
        'Priorizar contacto con estudiantes en riesgo alto y con pendientes actuales o entregas atrasadas.',
        'Validar manualmente en Canvas los casos con registro pendiente, incompleto o sin información de correo/SIS.',
        'No clasificar como abandono a estudiantes con curso finalizado o al día en entregables sin revisión adicional.',
        'Documentar en bitácora los intentos de contacto y las acciones de seguimiento realizadas por el asesor.',
    ]
    if summary is not None and not summary.empty and 'curso_general' in summary.columns:
        counts = _risk_counts(summary)
        if counts.get('Alto', 0) == 0:
            recs.insert(0, 'Mantener monitoreo preventivo: no se detectan casos de riesgo alto en el corte actual.')
    data = [['Recomendación operativa'], *[[r] for r in recs]]
    t = Table([[ _p(x[0], styles['SmallPremium'], 165)] for x in data], colWidths=[9.2*inch])
    t.setStyle(TableStyle([
        ('BACKGROUND',(0,0),(-1,0),colors.HexColor('#00A83B')),('TEXTCOLOR',(0,0),(-1,0),colors.white),
        ('GRID',(0,0),(-1,-1),0.25,colors.HexColor('#D9E2F3')),
        ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white, colors.HexColor('#F7F9FB')]),
        ('VALIGN',(0,0),(-1,-1),'TOP')
    ]))
    return t


def pdf_bytes(summary: pd.DataFrame, course_name: str, section_name: str, generated_by: str, analysis_date: str, logo_ave='assets/logo_ave.png', logo_uvg='assets/logo_uvg.png') -> bytes:
    """Informe ejecutivo premium AVE/UVG.

    Incluye portada, nombre del curso general, resumen ejecutivo, KPI, gráficas,
    resumen por curso, casos prioritarios y recomendaciones. Conserva la firma
    pública para que app.py no necesite cambios adicionales.
    """
    if summary is None:
        summary = pd.DataFrame()
    summary = summary.copy()
    styles = _premium_styles()
    curso_label = _course_label_from_summary(summary, course_name)
    total = len(summary)
    counts = _risk_counts(summary)
    cursos_generales = summary['curso_general'].nunique() if total and 'curso_general' in summary.columns else 1
    secciones = summary['curso_canvas'].nunique() if total and 'curso_canvas' in summary.columns else section_name
    finalizados = int(summary['estado_finalizacion'].astype(str).eq('Finalizado o al día por entregas').sum()) if total and 'estado_finalizacion' in summary.columns else 0
    no_reg = int(summary['clasificacion_registro'].astype(str).str.contains('no registrado|pendiente|incompleto', case=False, na=False).sum()) if total and 'clasificacion_registro' in summary.columns else 0
    avance_prom = f"{_num_mean(summary, 'porcentaje_avance'):.1f}%" if total else '0.0%'
    pendientes = int(_num_sum(summary, 'pendientes_actuales') if total and 'pendientes_actuales' in summary.columns else _num_sum(summary, 'pendientes')) if total else 0

    bio = BytesIO()
    doc = SimpleDocTemplate(bio, pagesize=landscape(letter), rightMargin=30, leftMargin=30, topMargin=34, bottomMargin=30)
    story = []

    # Portada compacta
    logo_row = []
    logo_row.append(Image(logo_ave, width=1.8*inch, height=0.82*inch) if Path(logo_ave).exists() else '')
    logo_row.append(Paragraph('Informe Ejecutivo de Seguimiento Académico', styles['CoverTitlePremium']))
    logo_row.append(Image(logo_uvg, width=0.95*inch, height=0.95*inch) if Path(logo_uvg).exists() else '')
    ht = Table([logo_row], colWidths=[2.1*inch, 6.15*inch, 1.35*inch])
    ht.setStyle(TableStyle([('VALIGN',(0,0),(-1,-1),'MIDDLE'), ('ALIGN',(0,0),(-1,-1),'CENTER')]))
    story += [ht, Spacer(1, 8)]
    story.append(Paragraph(f'<b>Curso general analizado:</b> {curso_label}', styles['CoverSubtitlePremium']))
    story.append(Paragraph(f'<b>Alcance:</b> {cursos_generales} curso(s) general(es) | {secciones} curso(s)/sección(es) Canvas | {total} estudiante(s) detectado(s)', styles['CoverSubtitlePremium']))
    story.append(Paragraph(f'<b>Fecha de corte:</b> {analysis_date} &nbsp;&nbsp; | &nbsp;&nbsp; <b>Generado por:</b> {generated_by or "No especificado"}', styles['CoverSubtitlePremium']))
    story += [Spacer(1, 12)]

    cards = [
        ('Estudiantes', total, '#172B85'),
        ('Riesgo alto', counts.get('Alto', 0), '#D64545'),
        ('Riesgo medio', counts.get('Medio', 0), '#F2C94C'),
        ('Riesgo bajo', counts.get('Bajo', 0), '#00A83B'),
        ('Finalizados / al día', finalizados, '#172B85'),
        ('Registros pendientes', no_reg, '#6c757d'),
        ('Avance promedio', avance_prom, '#172B85'),
        ('Pendientes actuales', pendientes, '#D64545'),
    ]
    story += [_kpi_cards(cards, styles), Spacer(1, 10)]
    story.append(Paragraph('Resumen general', styles['SectionPremium']))
    story.append(_executive_findings(summary, styles))
    story += [Spacer(1, 8)]
    story.append(Paragraph('Este reporte integra actividad de conexión, entregables, avance y clasificación de registros. La lectura principal debe concentrarse en el riesgo integral, porque ajusta el análisis para evitar priorizar indebidamente a estudiantes que ya finalizaron o se encuentran al día con sus entregas.', styles['BodyPremium']))

    story.append(PageBreak())
    story.append(Paragraph('Gráficos de sustento del análisis', styles['SectionPremium']))
    graph_table = Table([
        [_risk_pie_premium(summary), _course_risk_chart_premium(summary)],
        [_completion_chart_premium(summary), Paragraph('<b>Lectura sugerida:</b><br/>1. Revisar primero la proporción de riesgo alto y medio.<br/>2. Identificar el curso general con mayor concentración de alertas.<br/>3. Contrastar conexión con entregables para no penalizar estudiantes finalizados.<br/>4. Priorizar registros pendientes o incompletos para validación manual en Canvas.', styles['BodyPremium'])]
    ], colWidths=[4.25*inch, 5.15*inch], rowHeights=[2.35*inch, 2.25*inch])
    graph_table.setStyle(TableStyle([('VALIGN',(0,0),(-1,-1),'TOP'), ('GRID',(0,0),(-1,-1),0.25,colors.HexColor('#EEF2F7'))]))
    story += [graph_table, Spacer(1, 10)]

    if total and 'curso_general' in summary.columns:
        story.append(Paragraph('Resumen global por curso general', styles['SectionPremium']))
        tmp = summary.copy()
        for col in ['pendientes_actuales','pendientes','porcentaje_avance','horas_sin_actividad','atrasadas']:
            if col in tmp.columns:
                tmp[col] = pd.to_numeric(tmp[col], errors='coerce').fillna(0)
        if 'pendientes_actuales' not in tmp.columns:
            tmp['pendientes_actuales'] = tmp['pendientes'] if 'pendientes' in tmp.columns else 0
        if 'porcentaje_avance' not in tmp.columns:
            tmp['porcentaje_avance'] = 0
        if 'horas_sin_actividad' not in tmp.columns:
            tmp['horas_sin_actividad'] = 0
        g = tmp.groupby('curso_general').agg(
            estudiantes=('curso_general','count'),
            riesgo_alto=('riesgo_integral', lambda s: int((s == 'Alto').sum()) if 'riesgo_integral' in tmp.columns else 0),
            riesgo_medio=('riesgo_integral', lambda s: int((s == 'Medio').sum()) if 'riesgo_integral' in tmp.columns else 0),
            finalizados=('estado_finalizacion', lambda s: int((s == 'Finalizado o al día por entregas').sum()) if 'estado_finalizacion' in tmp.columns else 0),
            registros_pendientes=('clasificacion_registro', lambda s: int(s.astype(str).str.contains('no registrado|pendiente|incompleto', case=False, na=False).sum()) if 'clasificacion_registro' in tmp.columns else 0),
            pendientes_actuales=('pendientes_actuales','sum'),
            atrasadas=('atrasadas','sum') if 'atrasadas' in tmp.columns else ('pendientes_actuales','sum'),
            avance_prom=('porcentaje_avance','mean'),
            horas_sin_act_prom=('horas_sin_actividad','mean'),
        ).reset_index().sort_values(['riesgo_alto','riesgo_medio','pendientes_actuales'], ascending=False)
        g['avance_prom'] = pd.to_numeric(g['avance_prom'], errors='coerce').fillna(0).round(1).astype(str) + '%'
        g['horas_sin_act_prom'] = pd.to_numeric(g['horas_sin_act_prom'], errors='coerce').fillna(0).round(1).astype(str)
        for c in ['pendientes_actuales','atrasadas']:
            g[c] = pd.to_numeric(g[c], errors='coerce').fillna(0).astype(int)
        cols = ['curso_general','estudiantes','riesgo_alto','riesgo_medio','finalizados','registros_pendientes','pendientes_actuales','atrasadas','avance_prom','horas_sin_act_prom']
        heads = ['Curso general','Est.','Alto','Medio','Final./día','Reg. pend.','Pend. act.','Atr.','Avance','Hrs sin act. prom.']
        data = [heads]
        for _, r in g[cols].head(12).iterrows():
            data.append([_p(r['curso_general'], styles['TinyPremium'], 38), *[str(r[c]) for c in cols[1:]]])
        t = Table(data, repeatRows=1, colWidths=[2.1*inch,0.45*inch,0.45*inch,0.5*inch,0.62*inch,0.65*inch,0.65*inch,0.42*inch,0.58*inch,0.78*inch])
        t.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),colors.HexColor('#172B85')),('TEXTCOLOR',(0,0),(-1,0),colors.white),('GRID',(0,0),(-1,-1),0.22,colors.lightgrey),('FONTSIZE',(0,0),(-1,-1),6.2),('VALIGN',(0,0),(-1,-1),'TOP'),('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white, colors.HexColor('#F7F9FB')])]))
        story.append(t)

    story.append(PageBreak())
    story.append(Paragraph('Casos prioritarios y acciones recomendadas', styles['SectionPremium']))
    if summary.empty:
        story.append(Paragraph('No hay estudiantes para mostrar.', styles['BodyPremium']))
    else:
        view = summary.copy()
        # Columnas garantizadas
        needed = ['curso_general','estudiante','correo','clasificacion_registro','estado_finalizacion','horas_sin_actividad','pendientes_actuales','atrasadas','porcentaje_avance','porcentaje_avance_curso','riesgo_integral','accion_recomendada','puntaje_riesgo']
        for c in needed:
            if c not in view.columns:
                view[c] = '' if c not in ['puntaje_riesgo','horas_sin_actividad','pendientes_actuales','atrasadas','porcentaje_avance','porcentaje_avance_curso'] else 0
        for c in ['puntaje_riesgo','horas_sin_actividad','pendientes_actuales','atrasadas','porcentaje_avance','porcentaje_avance_curso']:
            view[c] = pd.to_numeric(view[c], errors='coerce').fillna(0)
        risk_ord = {'Alto': 0, 'Medio': 1, 'Bajo': 2}
        view['_risk_order'] = view['riesgo_integral'].map(risk_ord).fillna(3)
        mask = (view['riesgo_integral'].isin(['Alto','Medio'])) | (~view['clasificacion_registro'].astype(str).eq('Activo analizado'))
        view = view[mask].sort_values(['_risk_order','puntaje_riesgo','pendientes_actuales','horas_sin_actividad'], ascending=[True,False,False,False]).head(38)
        if view.empty:
            view = summary.head(25)
        heads = ['Curso','Estudiante','Correo','Registro','Estado','Hrs','Pend.','Atr.','Av.%','Riesgo','Acción recomendada']
        data = [heads]
        for _, r in view.iterrows():
            data.append([
                _p(r.get('curso_general',''), styles['TinyPremium'], 22),
                _p(r.get('estudiante',''), styles['TinyPremium'], 25),
                _p(r.get('correo',''), styles['TinyPremium'], 24),
                _p(r.get('clasificacion_registro',''), styles['TinyPremium'], 22),
                _p(r.get('estado_finalizacion',''), styles['TinyPremium'], 24),
                str(round(float(r.get('horas_sin_actividad',0)),1)),
                str(int(float(r.get('pendientes_actuales',0)))),
                str(int(float(r.get('atrasadas',0)))),
                str(round(float(r.get('porcentaje_avance',0)),1)),
                _p(r.get('riesgo_integral',''), styles['TinyPremium'], 8),
                _p(r.get('accion_recomendada',''), styles['TinyPremium'], 48),
            ])
        t = Table(data, repeatRows=1, colWidths=[0.92*inch,1.05*inch,1.0*inch,0.88*inch,0.9*inch,0.36*inch,0.36*inch,0.3*inch,0.34*inch,0.42*inch,2.05*inch])
        t.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),colors.HexColor('#6c757d')),('TEXTCOLOR',(0,0),(-1,0),colors.white),('GRID',(0,0),(-1,-1),0.2,colors.lightgrey),('FONTSIZE',(0,0),(-1,-1),5.8),('VALIGN',(0,0),(-1,-1),'TOP'),('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white, colors.HexColor('#F7F9FB')])]))
        story.append(t)
    story += [Spacer(1, 8), _recommendations(summary, styles), Spacer(1, 8)]
    story.append(Paragraph('Nota metodológica: la clasificación se basa en información disponible desde Canvas por medio del token autorizado. Los resultados deben utilizarse como insumo de asesoría académica, seguimiento preventivo y documentación operativa; no sustituyen la revisión pedagógica ni la comunicación directa con el estudiante.', styles['SmallPremium']))

    doc.build(story, onFirstPage=_executive_page, onLaterPages=_executive_page)
    return bio.getvalue()

# ===== PRO 4.0 FIX4b: ajuste final de diagramación sin traslapes =====
def _executive_findings(summary: pd.DataFrame, styles) -> Table:
    total = len(summary) if summary is not None else 0
    counts = _risk_counts(summary)
    high = counts.get('Alto', 0)
    med = counts.get('Medio', 0)
    pct_high = (high / total * 100) if total else 0
    finalizados = int(summary['estado_finalizacion'].astype(str).eq('Finalizado o al día por entregas').sum()) if total and 'estado_finalizacion' in summary.columns else 0
    no_reg = int(summary['clasificacion_registro'].astype(str).str.contains('no registrado|pendiente|incompleto', case=False, na=False).sum()) if total and 'clasificacion_registro' in summary.columns else 0
    pendientes = int(_num_sum(summary, 'pendientes_actuales') if total and 'pendientes_actuales' in summary.columns else _num_sum(summary, 'pendientes')) if total else 0
    course_top = 'No determinado'
    if total and 'curso_general' in summary.columns and 'riesgo_integral' in summary.columns:
        g = summary.groupby('curso_general')['riesgo_integral'].apply(lambda s: int((s == 'Alto').sum())).sort_values(ascending=False)
        if not g.empty and int(g.iloc[0]) > 0:
            course_top = f'{g.index[0]} ({int(g.iloc[0])} casos alto)'
    rows = [
        ['Hallazgo ejecutivo', 'Lectura operativa'],
        ['Riesgo alto', f'{high} de {total} estudiantes ({pct_high:.1f}%) requieren priorización inmediata.'],
        ['Riesgo medio', f'{med} estudiantes requieren prevención y seguimiento antes de convertirse en alerta crítica.'],
        ['Entregables', f'{finalizados} estudiantes aparecen finalizados o al día; este dato evita clasificarlos solo por baja conexión.'],
        ['Pendientes actuales', f'Se detectan {pendientes} pendientes actuales acumulados en los estudiantes analizados.'],
        ['Registros incompletos', f'{no_reg} estudiantes presentan información pendiente, no registrada o incompleta y deben verificarse en Canvas.'],
        ['Curso con mayor atención', course_top],
    ]
    data = [[rows[0][0], rows[0][1]]] + [[_p(c, styles['SmallPremium'], 36), _p(v, styles['SmallPremium'], 150)] for c, v in rows[1:]]
    t = Table(data, colWidths=[1.65*inch, 7.55*inch])
    t.setStyle(TableStyle([
        ('BACKGROUND',(0,0),(-1,0),colors.HexColor('#172B85')),('TEXTCOLOR',(0,0),(-1,0),colors.white),
        ('GRID',(0,0),(-1,-1),0.25,colors.HexColor('#D9E2F3')),
        ('VALIGN',(0,0),(-1,-1),'TOP'),
        ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white, colors.HexColor('#F7F9FB')]),
        ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold')
    ]))
    return t


def pdf_bytes(summary: pd.DataFrame, course_name: str, section_name: str, generated_by: str, analysis_date: str, logo_ave='assets/logo_ave.png', logo_uvg='assets/logo_uvg.png') -> bytes:
    if summary is None:
        summary = pd.DataFrame()
    summary = summary.copy()
    styles = _premium_styles()
    curso_label = _course_label_from_summary(summary, course_name)
    total = len(summary)
    counts = _risk_counts(summary)
    cursos_generales = summary['curso_general'].nunique() if total and 'curso_general' in summary.columns else 1
    secciones = summary['curso_canvas'].nunique() if total and 'curso_canvas' in summary.columns else section_name
    finalizados = int(summary['estado_finalizacion'].astype(str).eq('Finalizado o al día por entregas').sum()) if total and 'estado_finalizacion' in summary.columns else 0
    no_reg = int(summary['clasificacion_registro'].astype(str).str.contains('no registrado|pendiente|incompleto', case=False, na=False).sum()) if total and 'clasificacion_registro' in summary.columns else 0
    avance_prom = f"{_num_mean(summary, 'porcentaje_avance'):.1f}%" if total else '0.0%'
    pendientes = int(_num_sum(summary, 'pendientes_actuales') if total and 'pendientes_actuales' in summary.columns else _num_sum(summary, 'pendientes')) if total else 0

    bio = BytesIO()
    doc = SimpleDocTemplate(bio, pagesize=landscape(letter), rightMargin=30, leftMargin=30, topMargin=34, bottomMargin=30)
    story = []
    logo_row = [
        Image(logo_ave, width=1.8*inch, height=0.82*inch) if Path(logo_ave).exists() else '',
        Paragraph('Informe Ejecutivo de Seguimiento Académico', styles['CoverTitlePremium']),
        Image(logo_uvg, width=0.95*inch, height=0.95*inch) if Path(logo_uvg).exists() else ''
    ]
    ht = Table([logo_row], colWidths=[2.1*inch, 6.15*inch, 1.35*inch])
    ht.setStyle(TableStyle([('VALIGN',(0,0),(-1,-1),'MIDDLE'), ('ALIGN',(0,0),(-1,-1),'CENTER')]))
    story += [ht, Spacer(1, 8)]
    story.append(Paragraph(f'<b>Curso general analizado:</b> {curso_label}', styles['CoverSubtitlePremium']))
    story.append(Paragraph(f'<b>Alcance:</b> {cursos_generales} curso(s) general(es) | {secciones} curso(s)/sección(es) Canvas | {total} estudiante(s) detectado(s)', styles['CoverSubtitlePremium']))
    story.append(Paragraph(f'<b>Fecha de corte:</b> {analysis_date} &nbsp;&nbsp; | &nbsp;&nbsp; <b>Generado por:</b> {generated_by or "No especificado"}', styles['CoverSubtitlePremium']))
    story += [Spacer(1, 12)]
    cards = [
        ('Estudiantes', total, '#172B85'), ('Riesgo alto', counts.get('Alto', 0), '#D64545'),
        ('Riesgo medio', counts.get('Medio', 0), '#F2C94C'), ('Riesgo bajo', counts.get('Bajo', 0), '#00A83B'),
        ('Finalizados / al día', finalizados, '#172B85'), ('Registros pendientes', no_reg, '#6c757d'),
        ('Avance promedio', avance_prom, '#172B85'), ('Pendientes actuales', pendientes, '#D64545'),
    ]
    story += [_kpi_cards(cards, styles), Spacer(1, 10)]
    story.append(Paragraph('Resumen general', styles['SectionPremium']))
    story.append(_executive_findings(summary, styles))
    story += [Spacer(1, 8), Paragraph('Este reporte integra actividad de conexión, entregables, avance y clasificación de registros. La lectura principal debe concentrarse en el riesgo integral, porque ajusta el análisis para evitar priorizar indebidamente a estudiantes que ya finalizaron o se encuentran al día con sus entregas.', styles['BodyPremium'])]

    story.append(PageBreak())
    story.append(Paragraph('Gráficos de sustento del análisis', styles['SectionPremium']))
    graph_table = Table([[_risk_pie_premium(summary), _course_risk_chart_premium(summary)]], colWidths=[4.15*inch, 5.35*inch], rowHeights=[2.55*inch])
    graph_table.setStyle(TableStyle([('VALIGN',(0,0),(-1,-1),'TOP'), ('GRID',(0,0),(-1,-1),0.25,colors.HexColor('#EEF2F7'))]))
    story += [graph_table, Spacer(1, 8)]
    note = Table([[
        _completion_chart_premium(summary),
        Paragraph('<b>Lectura para toma de decisiones:</b><br/>• El riesgo alto exige intervención prioritaria.<br/>• El riesgo medio requiere prevención y revisión de entregables.<br/>• Los estudiantes finalizados o al día no deben tratarse como abandono solo por baja conexión.<br/>• Los registros pendientes o incompletos deben validarse manualmente en Canvas.', styles['BodyPremium'])
    ]], colWidths=[4.15*inch, 5.35*inch], rowHeights=[1.95*inch])
    note.setStyle(TableStyle([('VALIGN',(0,0),(-1,-1),'TOP'), ('GRID',(0,0),(-1,-1),0.25,colors.HexColor('#EEF2F7'))]))
    story.append(note)

    story.append(PageBreak())
    if total and 'curso_general' in summary.columns:
        story.append(Paragraph('Resumen global por curso general', styles['SectionPremium']))
        tmp = summary.copy()
        for col in ['pendientes_actuales','pendientes','porcentaje_avance','horas_sin_actividad','atrasadas']:
            if col in tmp.columns:
                tmp[col] = pd.to_numeric(tmp[col], errors='coerce').fillna(0)
        if 'pendientes_actuales' not in tmp.columns:
            tmp['pendientes_actuales'] = tmp['pendientes'] if 'pendientes' in tmp.columns else 0
        if 'porcentaje_avance' not in tmp.columns: tmp['porcentaje_avance'] = 0
        if 'horas_sin_actividad' not in tmp.columns: tmp['horas_sin_actividad'] = 0
        if 'atrasadas' not in tmp.columns: tmp['atrasadas'] = 0
        g = tmp.groupby('curso_general').agg(
            estudiantes=('curso_general','count'),
            riesgo_alto=('riesgo_integral', lambda s: int((s == 'Alto').sum()) if 'riesgo_integral' in tmp.columns else 0),
            riesgo_medio=('riesgo_integral', lambda s: int((s == 'Medio').sum()) if 'riesgo_integral' in tmp.columns else 0),
            finalizados=('estado_finalizacion', lambda s: int((s == 'Finalizado o al día por entregas').sum()) if 'estado_finalizacion' in tmp.columns else 0),
            registros_pendientes=('clasificacion_registro', lambda s: int(s.astype(str).str.contains('no registrado|pendiente|incompleto', case=False, na=False).sum()) if 'clasificacion_registro' in tmp.columns else 0),
            pendientes_actuales=('pendientes_actuales','sum'), atrasadas=('atrasadas','sum'), avance_prom=('porcentaje_avance','mean'), horas_sin_act_prom=('horas_sin_actividad','mean'),
        ).reset_index().sort_values(['riesgo_alto','riesgo_medio','pendientes_actuales'], ascending=False)
        g['avance_prom'] = pd.to_numeric(g['avance_prom'], errors='coerce').fillna(0).round(1).astype(str) + '%'
        g['horas_sin_act_prom'] = pd.to_numeric(g['horas_sin_act_prom'], errors='coerce').fillna(0).round(1).astype(str)
        for c in ['pendientes_actuales','atrasadas']:
            g[c] = pd.to_numeric(g[c], errors='coerce').fillna(0).astype(int)
        cols = ['curso_general','estudiantes','riesgo_alto','riesgo_medio','finalizados','registros_pendientes','pendientes_actuales','atrasadas','avance_prom','horas_sin_act_prom']
        heads = ['Curso general','Est.','Alto','Medio','Final./día','Reg. pend.','Pend. act.','Atr.','Avance','Hrs sin act. prom.']
        data = [heads]
        for _, r in g[cols].head(14).iterrows():
            data.append([_p(r['curso_general'], styles['TinyPremium'], 45), *[str(r[c]) for c in cols[1:]]])
        t = Table(data, repeatRows=1, colWidths=[2.1*inch,0.45*inch,0.45*inch,0.5*inch,0.62*inch,0.65*inch,0.65*inch,0.42*inch,0.58*inch,0.78*inch])
        t.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),colors.HexColor('#172B85')),('TEXTCOLOR',(0,0),(-1,0),colors.white),('GRID',(0,0),(-1,-1),0.22,colors.lightgrey),('FONTSIZE',(0,0),(-1,-1),6.2),('VALIGN',(0,0),(-1,-1),'TOP'),('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white, colors.HexColor('#F7F9FB')])]))
        story.append(t)
    story += [Spacer(1, 10), Paragraph('Recomendaciones ejecutivas', styles['SectionPremium']), _recommendations(summary, styles)]

    story.append(PageBreak())
    story.append(Paragraph('Casos prioritarios para seguimiento', styles['SectionPremium']))
    if summary.empty:
        story.append(Paragraph('No hay estudiantes para mostrar.', styles['BodyPremium']))
    else:
        view = summary.copy()
        needed = ['curso_general','estudiante','correo','clasificacion_registro','estado_finalizacion','horas_sin_actividad','pendientes_actuales','atrasadas','porcentaje_avance','riesgo_integral','accion_recomendada','puntaje_riesgo']
        for c in needed:
            if c not in view.columns: view[c] = '' if c not in ['puntaje_riesgo','horas_sin_actividad','pendientes_actuales','atrasadas','porcentaje_avance'] else 0
        for c in ['puntaje_riesgo','horas_sin_actividad','pendientes_actuales','atrasadas','porcentaje_avance']:
            view[c] = pd.to_numeric(view[c], errors='coerce').fillna(0)
        risk_ord = {'Alto': 0, 'Medio': 1, 'Bajo': 2}
        view['_risk_order'] = view['riesgo_integral'].map(risk_ord).fillna(3)
        mask = (view['riesgo_integral'].isin(['Alto','Medio'])) | (~view['clasificacion_registro'].astype(str).eq('Activo analizado'))
        view = view[mask].sort_values(['_risk_order','puntaje_riesgo','pendientes_actuales','horas_sin_actividad'], ascending=[True,False,False,False]).head(38)
        if view.empty: view = summary.head(25)
        heads = ['Curso','Estudiante','Correo','Registro','Estado','Hrs','Pend.','Atr.','Av.%','Riesgo','Acción recomendada']
        data = [heads]
        for _, r in view.iterrows():
            data.append([_p(r.get('curso_general',''), styles['TinyPremium'], 22), _p(r.get('estudiante',''), styles['TinyPremium'], 25), _p(r.get('correo',''), styles['TinyPremium'], 24), _p(r.get('clasificacion_registro',''), styles['TinyPremium'], 22), _p(r.get('estado_finalizacion',''), styles['TinyPremium'], 24), str(round(float(r.get('horas_sin_actividad',0)),1)), str(int(float(r.get('pendientes_actuales',0)))), str(int(float(r.get('atrasadas',0)))), str(round(float(r.get('porcentaje_avance',0)),1)), _p(r.get('riesgo_integral',''), styles['TinyPremium'], 8), _p(r.get('accion_recomendada',''), styles['TinyPremium'], 48)])
        t = Table(data, repeatRows=1, colWidths=[0.92*inch,1.05*inch,1.0*inch,0.88*inch,0.9*inch,0.36*inch,0.36*inch,0.3*inch,0.34*inch,0.42*inch,2.05*inch])
        t.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),colors.HexColor('#6c757d')),('TEXTCOLOR',(0,0),(-1,0),colors.white),('GRID',(0,0),(-1,-1),0.2,colors.lightgrey),('FONTSIZE',(0,0),(-1,-1),5.8),('VALIGN',(0,0),(-1,-1),'TOP'),('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white, colors.HexColor('#F7F9FB')])]))
        story.append(t)
    story += [Spacer(1, 8), Paragraph('Nota metodológica: la clasificación se basa en información disponible desde Canvas por medio del token autorizado. Los resultados deben utilizarse como insumo de asesoría académica, seguimiento preventivo y documentación operativa; no sustituyen la revisión pedagógica ni la comunicación directa con el estudiante.', styles['SmallPremium'])]
    doc.build(story, onFirstPage=_executive_page, onLaterPages=_executive_page)
    return bio.getvalue()

# Ajuste visual de gráficas para evitar recortes en reportes con nombres largos.
def _course_risk_chart_premium(summary: pd.DataFrame):
    from reportlab.graphics.shapes import Drawing, String
    from reportlab.graphics.charts.barcharts import VerticalBarChart
    d = Drawing(430, 190)
    d.add(String(12, 174, 'Riesgo alto y medio por curso general', fontSize=10.5, fillColor=colors.HexColor('#172B85'), fontName='Helvetica-Bold'))
    if summary is None or summary.empty or 'curso_general' not in summary.columns:
        d.add(String(45, 90, 'Sin datos por curso', fontSize=9, fillColor=colors.grey))
        return d
    tmp = summary.copy()
    if 'riesgo_integral' not in tmp.columns:
        tmp['riesgo_integral'] = 'Sin clasificar'
    g = tmp.groupby('curso_general').agg(
        alto=('riesgo_integral', lambda s: int((s == 'Alto').sum())),
        medio=('riesgo_integral', lambda s: int((s == 'Medio').sum())),
        total=('riesgo_integral', 'count')
    ).reset_index().sort_values(['alto','medio','total'], ascending=False).head(6)
    if g.empty:
        d.add(String(45, 90, 'Sin datos por curso', fontSize=9, fillColor=colors.grey))
        return d
    chart = VerticalBarChart()
    chart.x = 32; chart.y = 38; chart.height = 105; chart.width = 300
    chart.data = [g['alto'].astype(int).tolist(), g['medio'].astype(int).tolist()]
    names = [str(x).replace('Fundamentos de la Comunicación','Fund. Com.').replace('Iniciando tu Experiencia Virtual','Exp. Virtual').replace('Matemáticas','Matem.')[:12] for x in g['curso_general'].tolist()]
    chart.categoryAxis.categoryNames = names
    chart.categoryAxis.labels.angle = 25
    chart.categoryAxis.labels.fontSize = 5.8
    chart.valueAxis.valueMin = 0
    chart.valueAxis.labels.fontSize = 6
    chart.bars[0].fillColor = colors.HexColor('#D64545')
    chart.bars[1].fillColor = colors.HexColor('#F2C94C')
    d.add(chart)
    d.add(String(335, 140, 'Rojo: riesgo alto', fontSize=6.5, fillColor=colors.HexColor('#D64545')))
    d.add(String(335, 127, 'Amarillo: riesgo medio', fontSize=6.5, fillColor=colors.HexColor('#9A7400')))
    return d


def _completion_chart_premium(summary: pd.DataFrame):
    from reportlab.graphics.shapes import Drawing, String
    from reportlab.graphics.charts.barcharts import HorizontalBarChart
    d = Drawing(410, 165)
    d.add(String(12, 150, 'Entregables y estado de avance', fontSize=10.5, fillColor=colors.HexColor('#172B85'), fontName='Helvetica-Bold'))
    if summary is None or summary.empty:
        d.add(String(45, 80, 'Sin datos de entregables', fontSize=9, fillColor=colors.grey))
        return d
    final_col = summary['estado_finalizacion'].astype(str) if 'estado_finalizacion' in summary.columns else pd.Series(['No determinado']*len(summary))
    finalizados = int(final_col.eq('Finalizado o al día por entregas').sum())
    en_seguimiento = int(len(summary) - finalizados)
    no_reg = int(summary['clasificacion_registro'].astype(str).str.contains('no registrado|pendiente|incompleto', case=False, na=False).sum()) if 'clasificacion_registro' in summary.columns else 0
    pendientes = int(_num_sum(summary, 'pendientes_actuales') if 'pendientes_actuales' in summary.columns else _num_sum(summary, 'pendientes'))
    labels = ['Finalizados/al día', 'En seguimiento', 'Registros pendientes', 'Pendientes actuales']
    values = [finalizados, en_seguimiento, no_reg, pendientes]
    chart = HorizontalBarChart()
    chart.x = 130; chart.y = 35; chart.height = 88; chart.width = 230
    chart.data = [values]
    chart.categoryAxis.categoryNames = labels
    chart.categoryAxis.labels.fontSize = 6
    chart.valueAxis.labels.fontSize = 6
    chart.valueAxis.valueMin = 0
    chart.bars[0].fillColor = colors.HexColor('#172B85')
    d.add(chart)
    return d

# Reducción final de gráfico de entregables para respetar el ancho de la celda PDF.
def _completion_chart_premium(summary: pd.DataFrame):
    from reportlab.graphics.shapes import Drawing, String
    from reportlab.graphics.charts.barcharts import HorizontalBarChart
    d = Drawing(290, 165)
    d.add(String(10, 150, 'Entregables y estado de avance', fontSize=10.5, fillColor=colors.HexColor('#172B85'), fontName='Helvetica-Bold'))
    if summary is None or summary.empty:
        d.add(String(35, 80, 'Sin datos de entregables', fontSize=9, fillColor=colors.grey))
        return d
    final_col = summary['estado_finalizacion'].astype(str) if 'estado_finalizacion' in summary.columns else pd.Series(['No determinado']*len(summary))
    finalizados = int(final_col.eq('Finalizado o al día por entregas').sum())
    en_seguimiento = int(len(summary) - finalizados)
    no_reg = int(summary['clasificacion_registro'].astype(str).str.contains('no registrado|pendiente|incompleto', case=False, na=False).sum()) if 'clasificacion_registro' in summary.columns else 0
    pendientes = int(_num_sum(summary, 'pendientes_actuales') if 'pendientes_actuales' in summary.columns else _num_sum(summary, 'pendientes'))
    labels = ['Finalizados/al día', 'En seguimiento', 'Registros pendientes', 'Pendientes actuales']
    values = [finalizados, en_seguimiento, no_reg, pendientes]
    chart = HorizontalBarChart()
    chart.x = 112; chart.y = 35; chart.height = 88; chart.width = 145
    chart.data = [values]
    chart.categoryAxis.categoryNames = labels
    chart.categoryAxis.labels.fontSize = 5.7
    chart.valueAxis.labels.fontSize = 6
    chart.valueAxis.valueMin = 0
    chart.bars[0].fillColor = colors.HexColor('#172B85')
    d.add(chart)
    return d
